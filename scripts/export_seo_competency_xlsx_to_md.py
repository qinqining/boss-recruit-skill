#!/usr/bin/env python3
"""
可选：若本地存在「面试方案生成任务/SEO运营岗位素质项.xlsx」，则覆盖写入
「面试方案生成任务/岗位素质项/」下 01、02 两个 md。不生成 README，不写「导出自」。
仓库根目录（Windows 推荐 py）：py scripts/export_seo_competency_xlsx_to_md.py
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from e

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "面试方案生成任务" / "SEO运营岗位素质项.xlsx"
OUT_DIR = ROOT / "面试方案生成任务" / "岗位素质项"


def cell_red(cell) -> bool:
    if not cell.font or not cell.font.color:
        return False
    rgb = getattr(cell.font.color, "rgb", None)
    if not rgb:
        return False
    rgs = str(rgb).upper()
    return "FF0000" in rgs or rgs.endswith("FF0000")


def norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def main() -> int:
    if not XLSX.is_file():
        print(f"未找到 xlsx，跳过：{XLSX}", file=sys.stderr)
        return 0

    wb = openpyxl.load_workbook(XLSX, data_only=False)
    if "SEO运营" not in wb.sheetnames:
        print("工作表中无「SEO运营」：", wb.sheetnames, file=sys.stderr)
        return 1
    ws = wb["SEO运营"]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    lines: list[str] = ["# 岗位与 JD 门槛\n"]
    r1b = norm(ws.cell(1, 2).value)
    r2a, r2b = norm(ws.cell(2, 1).value), norm(ws.cell(2, 2).value)
    r3a, r3b = norm(ws.cell(3, 1).value), norm(ws.cell(3, 2).value)
    lines.append(f"## 岗位名称\n\n{r1b}\n")
    lines.append(f"## {r2a or '岗位职责'}\n\n{r2b}\n")
    lines.append(f"## {r3a or '门槛（冰山上）'}\n\n{r3b}\n")
    (OUT_DIR / "01_岗位与JD门槛.md").write_text("\n".join(lines), encoding="utf-8")

    lines2: list[str] = ["# 素质项、分档行为锚点与面试母题\n\n"]
    hdr_c = norm(ws.cell(4, 3).value)
    hdr_d = norm(ws.cell(4, 4).value)
    hdr_e = norm(ws.cell(4, 5).value)
    hdr_f = norm(ws.cell(4, 6).value)
    hdr_g = norm(ws.cell(4, 7).value)
    hdr_h = norm(ws.cell(4, 8).value)

    for ri in range(5, 10):
        comp = re.sub(r"\s+", " ", norm(ws.cell(ri, 2).value)).strip()
        trait = norm(ws.cell(ri, 3).value)
        d = norm(ws.cell(ri, 4).value)
        e = norm(ws.cell(ri, 5).value)
        f = norm(ws.cell(ri, 6).value)
        g = norm(ws.cell(ri, 7).value)
        h_cell = ws.cell(ri, 8)
        h_text = norm(h_cell.value)
        red = cell_red(h_cell)

        lines2.append(f"\n## {comp}\n")
        if trait:
            lines2.append(f"- **{hdr_c or '对应的性格特质'}**：{trait}\n")
        lines2.append(f"### {hdr_d or '0–3分'}\n\n{d}\n")
        lines2.append(f"### {hdr_e or '4–6分'}\n\n{e}\n")
        lines2.append(f"### {hdr_f or '7–8分'}\n\n{f}\n")
        lines2.append(f"### {hdr_g or '9–10分'}\n\n{g}\n")
        red_note = (
            "（**高优先级母题**，生成面试方案须覆盖其考察意图）" if red else ""
        )
        lines2.append(f"### {hdr_h or '对应的面试问题'}{red_note}\n\n{h_text}\n")

    (OUT_DIR / "02_素质分档与面试母题.md").write_text("".join(lines2), encoding="utf-8")
    print(f"已写入：{OUT_DIR.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
